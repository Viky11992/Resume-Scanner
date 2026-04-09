"""
Ranker Module

Ranks, shortlists, and exports candidate matching results to Excel and PDF formats.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from fpdf import FPDF


def rank_candidates(batch_results):
    """
    Sort candidates by their matching scores in descending order.

    Args:
        batch_results (list): List of dicts from batch_match(), each containing:
            - name (str): Candidate name.
            - score (float): Similarity score.
            - skills_matched (list): Matched skills.
            - skills_missing (list): Missing skills.

    Returns:
        list: Sorted list of candidate results (highest score first).

    Raises:
        ValueError: If batch_results is empty or invalid.
        Exception: If there is an error during ranking.
    """
    try:
        if not batch_results:
            raise ValueError("Batch results list is empty")

        ranked = sorted(batch_results, key=lambda x: x.get("score", 0), reverse=True)
        return ranked

    except ValueError:
        raise
    except Exception as e:
        raise Exception(f"Error ranking candidates: {str(e)}")


def shortlist_candidates(ranked_list, top_n=10):
    """
    Return the top N candidates from a ranked list.

    Args:
        ranked_list (list): Sorted list of candidate results.
        top_n (int): Number of top candidates to return. Defaults to 10.

    Returns:
        list: Top N candidates.

    Raises:
        ValueError: If ranked_list is empty or top_n is invalid.
        Exception: If there is an error during shortlisting.
    """
    try:
        if not ranked_list:
            raise ValueError("Ranked list is empty")
        if not isinstance(top_n, int) or top_n < 1:
            raise ValueError("top_n must be a positive integer")

        return ranked_list[:top_n]

    except ValueError:
        raise
    except Exception as e:
        raise Exception(f"Error shortlisting candidates: {str(e)}")


def export_to_excel(ranked_list, output_path):
    """
    Export ranked candidate results to an Excel file.

    Columns: Rank, Name, Score %, Skills Matched, Skills Missing

    Args:
        ranked_list (list): Sorted list of candidate results.
        output_path (str): File path for the output Excel file.

    Raises:
        ValueError: If ranked_list is empty or output_path is invalid.
        Exception: If there is an error during export.
    """
    try:
        if not ranked_list:
            raise ValueError("Ranked list is empty")
        if not output_path:
            raise ValueError("Output path is required")

        wb = Workbook()
        ws = wb.active
        ws.title = "Candidate Rankings"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Headers
        headers = ["Rank", "Name", "Score %", "Skills Matched", "Skills Missing"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Data rows
        for row_idx, candidate in enumerate(ranked_list, 2):
            rank = row_idx - 1
            name = candidate.get("name", "Unknown")
            score_pct = round(candidate.get("score", 0) * 100, 2)
            skills_matched = ", ".join(candidate.get("skills_matched", []))
            skills_missing = ", ".join(candidate.get("skills_missing", []))

            row_data = [rank, name, score_pct, skills_matched, skills_missing]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border

            # Color code score cells
            score_cell = ws.cell(row=row_idx, column=3)
            if score_pct >= 70:
                score_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif score_pct >= 50:
                score_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                score_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Adjust column widths
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 40
        ws.column_dimensions["E"].width = 40

        # Freeze header row
        ws.freeze_panes = "A2"

        wb.save(output_path)

    except ValueError:
        raise
    except Exception as e:
        raise Exception(f"Error exporting to Excel: {str(e)}")


def export_to_pdf(ranked_list, output_path):
    """
    Export ranked candidate results to a PDF report.

    Generates a clean, formatted PDF with candidate rankings table.

    Args:
        ranked_list (list): Sorted list of candidate results.
        output_path (str): File path for the output PDF file.

    Raises:
        ValueError: If ranked_list is empty or output_path is invalid.
        Exception: If there is an error during export.
    """
    try:
        if not ranked_list:
            raise ValueError("Ranked list is empty")
        if not output_path:
            raise ValueError("Output path is required")

        class PDF(FPDF):
            def header(self):
                self.set_font("Helvetica", "B", 16)
                self.set_text_color(47, 84, 150)
                self.cell(0, 10, "Resume Screener - Candidate Rankings", align="C")
                self.ln(12)
                self.set_draw_color(47, 84, 150)
                self.line(10, self.get_y(), 200, self.get_y())
                self.ln(5)

            def footer(self):
                self.set_y(-15)
                self.set_font("Helvetica", "I", 8)
                self.set_text_color(128, 128, 128)
                self.cell(0, 10, f"Page {self.page_no()}/{{nb}}", align="C")

        pdf = PDF()
        pdf.alias_nb_pages()
        pdf.add_page()

        # Summary
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(0, 0, 0)
        total_candidates = len(ranked_list)
        avg_score = round(
            sum(c.get("score", 0) for c in ranked_list) / total_candidates * 100, 2
        ) if total_candidates > 0 else 0
        pdf.cell(0, 8, f"Total Candidates: {total_candidates}")
        pdf.ln(8)
        pdf.cell(0, 8, f"Average Score: {avg_score}%")
        pdf.ln(8)

        # Table headers
        col_widths = [15, 40, 22, 56, 57]
        headers = ["Rank", "Name", "Score %", "Skills Matched", "Skills Missing"]

        pdf.set_font("Helvetica", "B", 9)
        pdf.set_fill_color(47, 84, 150)
        pdf.set_text_color(255, 255, 255)

        for i, header in enumerate(headers):
            pdf.cell(col_widths[i], 10, header, border=1, fill=True, align="C")
        pdf.ln()

        # Table data
        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(0, 0, 0)

        fill = False
        for idx, candidate in enumerate(ranked_list):
            rank = idx + 1
            name = candidate.get("name", "Unknown")
            score_pct = round(candidate.get("score", 0) * 100, 1)
            skills_matched = ", ".join(candidate.get("skills_matched", []))
            skills_missing = ", ".join(candidate.get("skills_missing", []))

            # Check if we need a new page
            if pdf.get_y() > 250:
                pdf.add_page()
                # Re-draw headers
                pdf.set_font("Helvetica", "B", 9)
                pdf.set_fill_color(47, 84, 150)
                pdf.set_text_color(255, 255, 255)
                for i, header in enumerate(headers):
                    pdf.cell(col_widths[i], 10, header, border=1, fill=True, align="C")
                pdf.ln()
                pdf.set_font("Helvetica", "", 8)
                pdf.set_text_color(0, 0, 0)

            # Alternate row colors
            if fill:
                pdf.set_fill_color(240, 240, 240)
            else:
                pdf.set_fill_color(255, 255, 255)

            row_height = 10
            max_lines = max(
                len(skills_matched) // 50 + 1,
                len(skills_missing) // 50 + 1,
                1,
            )
            row_height = max(10, max_lines * 8)

            y_start = pdf.get_y()
            x_start = pdf.get_x()

            # Draw cells
            pdf.set_xy(x_start, y_start)
            pdf.cell(col_widths[0], row_height, str(rank), border=1, fill=True, align="C")
            pdf.set_xy(x_start + col_widths[0], y_start)
            pdf.cell(col_widths[1], row_height, name, border=1, fill=True, align="C")
            pdf.set_xy(x_start + col_widths[0] + col_widths[1], y_start)
            pdf.cell(col_widths[2], row_height, f"{score_pct}%", border=1, fill=True, align="C")
            pdf.set_xy(x_start + col_widths[0] + col_widths[1] + col_widths[2], y_start)
            pdf.multi_cell(col_widths[3], row_height / max_lines, skills_matched, border=1, fill=True)
            pdf.set_xy(x_start + sum(col_widths[:4]), y_start)
            pdf.multi_cell(col_widths[4], row_height / max_lines, skills_missing, border=1, fill=True)

            pdf.set_xy(x_start, y_start + row_height)
            fill = not fill

        pdf.output(output_path)

    except ValueError:
        raise
    except Exception as e:
        raise Exception(f"Error exporting to PDF: {str(e)}")
